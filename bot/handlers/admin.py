import asyncio
import os

from aiogram import Router, types
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import FSInputFile, KeyboardButton, ReplyKeyboardMarkup
from handlers.logic import list_admin_info
from openpyxl import Workbook, load_workbook

router = Router()
# якась параша завтра пвирішу це питання


class Form(StatesGroup):
    ip = State()


@router.message(lambda message: message.text == "cmd_admin")
async def cmd_admin(message: types.Message, state: FSMContext):
    if message.from_user.username == "ds0903":
        kb = [
            [KeyboardButton(text="Всі активні ip користувачів")],
            [KeyboardButton(text="Всі користувачі в базі")],
            [KeyboardButton(text="Відвідувачі")],
            [KeyboardButton(text="Змінити значення")],
            [KeyboardButton(text="Забанить користувача")],
            [KeyboardButton(text="Видалити значення")],
        ]
        keyboard = ReplyKeyboardMarkup(keyboard=kb, resize_keyboard=True)
        await message.answer("Ти потрапив в секретне меню", reply_markup=keyboard)
    else:
        await message.answer("Доступ заборонено !")

    @router.message(lambda message: message.text == "Всі активні ip користувачів")
    async def cmd_all_data(message: types.Message, state: FSMContext):
        await message.answer("Всі активні ip які тільки є в базі")
        data = await list_admin_info(status="1")
        try:
            wb = load_workbook("all_user_active_ip.xlsx")
        except FileNotFoundError:
            wb = Workbook()

        sheet = wb.active
        sheet.append(
            [
                "ID",
                "User ID",
                "IP",
                "Description",
                "First Name",
                "Last Name",
                "Username",
                "Language Code",
                "Is Premium",
                "Is Active",
            ]
        )

        for i in data:
            (
                id,
                user_id,
                ip,
                description,
                first_name,
                last_name,
                username,
                language_code,
                is_premium,
                is_active,
            ) = i

            sheet.append(
                [
                    id,
                    user_id,
                    ip,
                    description,
                    first_name,
                    last_name,
                    username,
                    language_code,
                    is_premium,
                    is_active,
                ]
            )
            # await message.answer(f"Номер ip:{id}\nid користувача:{user_id}\nip користувача:{ip}\nОпис:{description}\nім'я: {first_name}\nПрізвище: {last_name}\nНікнейм: {username}\nМова: {language_code}\nЧи преміум?: {is_premium}\nчи активний: {is_active}")
        wb.save("all_user_active_ip.xlsx")

        file = FSInputFile("all_user_active_ip.xlsx")
        await message.answer_document(file)
        os.remove("all_user_active_ip.xlsx")

    @router.message(lambda message: message.text == "Всі користувачі в базі")
    async def cmd_all_users(message: types.Message, state: FSMContext):
        await message.answer("Всі користувачі які тільки є в базі")
        await message.answer("Відвідувачі")
        data1 = await list_admin_info(status="2")
        try:
            wb = load_workbook("all_users.xlsx")
        except FileNotFoundError:
            wb = Workbook()

        sheet = wb.active
        sheet.append(
            [
                "id",
                "user_id",
                "ip",
                "description",
                "first_name",
                "last_name",
                "username",
                "language_code",
                "is_premium",
            ]
        )

        for i in data1:
            (
                id,
                user_id,
                ip,
                description,
                first_name,
                last_name,
                username,
                language_code,
                is_premium,
            ) = i

            sheet.append(
                [
                    id,
                    user_id,
                    ip,
                    description,
                    first_name,
                    last_name,
                    username,
                    language_code,
                    is_premium,
                ]
            )
        wb.save("all_users.xlsx")

        file = FSInputFile("all_users.xlsx")
        await message.answer_document(file)
        os.remove("all_users.xlsx")

    @router.message(lambda message: message.text == "Відвідувачі")
    async def gosti(message: types.Message, state: FSMContext):
        await message.answer("Відвідувачі")
        data = await list_admin_info(status="3")
        try:
            wb = load_workbook("gosti.xlsx")
        except FileNotFoundError:
            wb = Workbook()

        sheet = wb.active
        sheet.append(
            [
                "user_id",
                "is_bot",
                "first_name",
                "last_name",
                "username",
                "language_code",
                "is_premium",
                "added_to_attachment_menu",
                "can_join_groups",
                "can_read_all_group_messages",
                "supports_inline_queries",
                "can_connect_to_business",
                "current_time",
            ]
        )

        for i in data:
            (
                id,
                user_id,
                is_bot,
                first_name,
                last_name,
                username,
                language_code,
                is_premium,
                added_to_attachment_menu,
                can_join_groups,
                can_read_all_group_messages,
                supports_inline_queries,
                can_connect_to_business,
                current_time,
            ) = i

            sheet.append(
                [
                    user_id,
                    is_bot,
                    first_name,
                    last_name,
                    username,
                    language_code,
                    is_premium,
                    added_to_attachment_menu,
                    can_join_groups,
                    can_read_all_group_messages,
                    supports_inline_queries,
                    can_connect_to_business,
                    current_time,
                ]
            )
        wb.save("gosti.xlsx")

        file = FSInputFile("gosti.xlsx")
        await message.answer_document(file)
        os.remove("gosti.xlsx")

    # Поки в роботі
    # @router.message(lambda message: message.text == "Забанить користувача")
    # async def cmd_ban_user(message: types.Message, state: FSMContext):
    #     await message.answer("Вибери користувача який буде заблокований")

    # @router.message(lambda message: message.text == "Видалити значення")
    # async def cmd_delete_data(message: types.Message, state: FSMContext):
    #     await message.answer("Вибери значення як потрібно видалити")

    # @router.message(lambda message: message.text == "Змінити значення")
    # async def cmd_change_data(message: types.Message, state: FSMContext):
    #     await message.answer("Вибери значення яке буде змінино")
